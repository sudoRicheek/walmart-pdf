"""
Parse all Walmart order XLSX files from ./data and populate a SQLite database.

Schema
------
orders:  order_number, order_date, shipping_address, payment_last4,
         subtotal, delivery_charges, tax, tip, order_total
items:   order_id (FK), product_name, quantity, price, delivery_status, product_link
splits:  item_id (FK), roommate (text), checked (0/1)
"""

import glob
import os
import re
import sqlite3
from datetime import datetime

import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), "walmart.db")
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

ROOMMATES = ["Roommate 1", "Roommate 2", "Roommate 3"]

# ── helpers ─────────────────────────────────────────────────────────
META_KEYS = {
    "Order Number",
    "Order Date",
    "Shipping Address",
    "Payment Method",
    "Subtotal",
    "Delivery Charges",
    "Tax",
    "Tip",
    "Order Total",
}


def _extract_last4(payment_str: str) -> str:
    """'Ending in 3392' → '3392'"""
    if not payment_str:
        return ""
    m = re.search(r"(\d{4})\s*$", str(payment_str))
    return m.group(1) if m else str(payment_str)


def _parse_date(date_str: str) -> str:
    """'Feb 06, 2026' → '2026-02-06'"""
    if not date_str:
        return ""
    try:
        return datetime.strptime(str(date_str).strip(), "%b %d, %Y").strftime("%Y-%m-%d")
    except ValueError:
        return str(date_str)


# ── database setup ──────────────────────────────────────────────────
def init_db(conn: sqlite3.Connection):
    c = conn.cursor()
    c.executescript(
        """
        DROP TABLE IF EXISTS splits;
        DROP TABLE IF EXISTS items;
        DROP TABLE IF EXISTS orders;

        CREATE TABLE orders (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number    TEXT UNIQUE NOT NULL,
            order_date      TEXT,          -- ISO 8601
            shipping_address TEXT,
            payment_last4   TEXT,
            subtotal        REAL DEFAULT 0,
            delivery_charges REAL DEFAULT 0,
            tax             REAL DEFAULT 0,
            tip             REAL DEFAULT 0,
            order_total     REAL DEFAULT 0
        );

        CREATE TABLE items (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id        INTEGER NOT NULL REFERENCES orders(id),
            product_name    TEXT NOT NULL,
            quantity        INTEGER DEFAULT 1,
            price           REAL DEFAULT 0,
            delivery_status TEXT,
            product_link    TEXT
        );

        CREATE TABLE splits (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id   INTEGER NOT NULL REFERENCES items(id),
            roommate  TEXT NOT NULL,
            checked   INTEGER DEFAULT 1,   -- 1 = included in split
            UNIQUE(item_id, roommate)
        );
        """
    )
    # roommates table is NOT dropped so custom names survive re-imports
    c.execute("""
        CREATE TABLE IF NOT EXISTS roommates (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT UNIQUE NOT NULL,
            sort_order INTEGER DEFAULT 0
        )
    """)
    conn.commit()


# ── xlsx parsing ────────────────────────────────────────────────────
def parse_xlsx(path: str):
    """Return (meta_dict, list_of_item_dicts) from one xlsx file."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Order Invoice"]

    meta: dict = {}
    items: list[dict] = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        first_val = row[0].value
        if first_val is None or first_val == "Product Name":
            continue
        if first_val in META_KEYS:
            meta[first_val] = ws.cell(row=row[0].row, column=row[0].column + 1).value
        else:
            # It's a product row
            items.append(
                {
                    "product_name": row[0].value,
                    "quantity": row[1].value or 0,
                    "price": row[2].value or 0.0,
                    "delivery_status": row[3].value or "",
                    "product_link": row[4].value or "",
                }
            )
    return meta, items


# ── main ────────────────────────────────────────────────────────────
def load_all():
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    c = conn.cursor()

    # Seed roommates table if empty (preserves any renames done via UI)
    rm_count = c.execute("SELECT COUNT(*) FROM roommates").fetchone()[0]
    if rm_count == 0:
        for i, name in enumerate(ROOMMATES):
            c.execute(
                "INSERT OR IGNORE INTO roommates (name, sort_order) VALUES (?, ?)",
                (name, i),
            )
        conn.commit()
    db_roommates = [
        r[0] for r in c.execute(
            "SELECT name FROM roommates ORDER BY sort_order, id"
        ).fetchall()
    ]

    files = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx")))
    print(f"Found {len(files)} xlsx files in {DATA_DIR}")
    print(f"Splitting among: {db_roommates}")

    for fpath in files:
        meta, items = parse_xlsx(fpath)
        order_number = meta.get("Order Number", os.path.basename(fpath))
        order_date = _parse_date(meta.get("Order Date", ""))
        shipping_address = meta.get("Shipping Address", "")
        payment_last4 = _extract_last4(meta.get("Payment Method", ""))
        subtotal = float(meta.get("Subtotal", 0) or 0)
        delivery_charges = float(meta.get("Delivery Charges", 0) or 0)
        tax = float(meta.get("Tax", 0) or 0)
        tip = float(meta.get("Tip", 0) or 0)
        order_total = float(meta.get("Order Total", 0) or 0)

        c.execute(
            """INSERT OR IGNORE INTO orders
               (order_number, order_date, shipping_address, payment_last4,
                subtotal, delivery_charges, tax, tip, order_total)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (
                order_number,
                order_date,
                shipping_address,
                payment_last4,
                subtotal,
                delivery_charges,
                tax,
                tip,
                order_total,
            ),
        )
        order_id = c.lastrowid

        for item in items:
            c.execute(
                """INSERT INTO items (order_id, product_name, quantity, price,
                                     delivery_status, product_link)
                   VALUES (?,?,?,?,?,?)""",
                (
                    order_id,
                    item["product_name"],
                    item["quantity"],
                    item["price"],
                    item["delivery_status"],
                    item["product_link"],
                ),
            )
            item_id = c.lastrowid
            # Default: all roommates checked
            for rm in db_roommates:
                c.execute(
                    "INSERT INTO splits (item_id, roommate, checked) VALUES (?,?,?)",
                    (item_id, rm, 1),
                )

        print(f"  ✓ {order_number}  {order_date}  card:{payment_last4}  items:{len(items)}")

    conn.commit()
    total_orders = c.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
    total_items = c.execute("SELECT COUNT(*) FROM items").fetchone()[0]
    print(f"\nDone — {total_orders} orders, {total_items} items loaded into {DB_PATH}")
    conn.close()


if __name__ == "__main__":
    load_all()
